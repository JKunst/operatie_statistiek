"""
Dataset generator voor Operatie Sterfgeval.
Genereert een unieke .xlsx dataset per leerlingnummer (6 cijfers).
"""

import math
import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuratie ─────────────────────────────────────────────────────────────

ARTSEN = [
    {"naam": "Dr. E. Bakker",   "specialisme": "Interne geneeskunde", "code": "BAK"},
    {"naam": "Dr. R. Jansen",   "specialisme": "Interne geneeskunde", "code": "JAN"},
    {"naam": "Dr. S. Vermeer",  "specialisme": "Interne geneeskunde", "code": "VER"},
    {"naam": "Dr. M. de Vries", "specialisme": "Interne geneeskunde", "code": "DVR"},
]

VERPLEEGKUNDIGEN = [
    "A. Smits", "B. Peters", "C. Hendriks",
    "D. Wolters", "E. Claassen", "F. Mulder",
]

DIENSTEN = [
    {"naam": "Ochtenddienst", "start": "07:00", "eind": "15:00"},
    {"naam": "Middagdienst",  "start": "15:00", "eind": "23:00"},
    {"naam": "Nachtdienst",   "start": "23:00", "eind": "07:00"},
]

WEKEN = 12
START_DATUM = date(2024, 1, 8)
DAGNAMES = ["Maandag","Dinsdag","Woensdag","Donderdag","Vrijdag","Zaterdag","Zondag"]

# Sterftekansen
NORM_MIN, NORM_MAX = 0.014, 0.026
VERD_MIN, VERD_MAX = 0.058, 0.082

# Kleuren
K_DONKER  = "1A1410"
K_ROOD    = "8B1A1A"
K_KOPTEKST = "2C3E50"
K_GEEL    = "FFF3CD"
K_ROOD_RIJ = "FDECEA"
K_GROEN_RIJ = "EBF5EB"
K_GRIJS   = "F5F5F5"
K_LIJN    = "DDDDDD"
K_WIT     = "FFFFFF"
K_PAPER   = "FAF6EE"
K_GROEN   = "2D5016"


# ── Seeded random ─────────────────────────────────────────────────────────────

def _sr(seed, idx):
    x = math.sin(seed * 9301 + idx * 49297 + 233) * 10000
    return x - math.floor(x)

def _hash(ln_str):
    h = 0
    for ch in str(ln_str):
        h = ((h << 5) - h) + ord(ch)
        h &= 0xFFFFFFFF
    return h


# ── Stijlhelpers ──────────────────────────────────────────────────────────────

def _fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _border():
    s = Side(style="thin", color=K_LIJN)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _style_hdr(cell, bg=K_KOPTEKST, fg=K_WIT, size=10):
    cell.font = _font(True, size, fg)
    cell.fill = _fill(bg)
    cell.alignment = _center()
    cell.border = _border()

def _style_data(cell, bg=K_WIT, bold=False, align="left", fmt=None, size=9):
    cell.font = _font(bold, size)
    cell.fill = _fill(bg)
    cell.alignment = _center() if align == "center" else _left()
    cell.border = _border()
    if fmt:
        cell.number_format = fmt


# ── Data genereren ────────────────────────────────────────────────────────────

def _genereer_rijen(leerlingnummer):
    seed = _hash(str(leerlingnummer))
    verdachte_idx = seed % len(ARTSEN)
    rijen = []
    ri = 0  # row index voor seeded random

    for dag in range(WEKEN * 7):
        datum = START_DATUM + timedelta(days=dag)
        is_weekend = datum.weekday() >= 5
        min_pat = 7 if is_weekend else 10
        max_pat = 14 if is_weekend else 20

        for d_idx, dienst in enumerate(DIENSTEN):
            arts_idx = int(_sr(seed, ri) * len(ARTSEN)) % len(ARTSEN)
            ri += 1
            # Rotatiesysteem: arts wisselt per dag/dienst
            arts_idx = (dag * 3 + d_idx + arts_idx) % len(ARTSEN)
            arts = ARTSEN[arts_idx]
            is_verdacht = (arts_idx == verdachte_idx)

            vplk_idx = int(_sr(seed, ri) * len(VERPLEEGKUNDIGEN))
            ri += 1
            vplk = VERPLEEGKUNDIGEN[vplk_idx]

            n_pat = min_pat + int(_sr(seed, ri) * (max_pat - min_pat + 1))
            ri += 1

            if is_verdacht:
                kans = VERD_MIN + _sr(seed, ri) * (VERD_MAX - VERD_MIN)
            else:
                kans = NORM_MIN + _sr(seed, ri) * (NORM_MAX - NORM_MIN)
            ri += 1

            overlijden = sum(1 for p in range(n_pat) if _sr(seed, ri + p) < kans)
            ri += n_pat

            ontslag_kans = 0.15 if is_weekend else 0.22
            max_ont = max(0, n_pat - overlijden)
            ontslagen = max(0, min(max_ont,
                round(n_pat * ontslag_kans + _sr(seed, ri) * 2 - 1)))
            ri += 1
            stabiel = n_pat - overlijden - ontslagen

            rijen.append({
                "datum": datum,
                "dag": DAGNAMES[datum.weekday()],
                "weeknr": datum.isocalendar()[1],
                "dienst": dienst["naam"],
                "start": dienst["start"],
                "arts_naam": arts["naam"],
                "arts_code": arts["code"],
                "vplk": vplk,
                "n_pat": n_pat,
                "ontslagen": ontslagen,
                "stabiel": stabiel,
                "overlijden": overlijden,
                "kans_pct": round(kans * 100, 2),
            })

    return rijen, verdachte_idx


# ── Tabblad 1: Opdracht ───────────────────────────────────────────────────────

def _maak_opdracht(wb, leerlingnummer, verdachte_idx):
    ws = wb.active
    ws.title = "Opdracht"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 66
    ws.column_dimensions["D"].width = 2

    # Titelbalk
    ws.row_dimensions[1].height = 8
    ws.merge_cells("B2:C3")
    c = ws["B2"]
    c.value = "OPERATIE STERFGEVAL"
    c.font = Font(name="Arial", bold=True, size=24, color=K_ROOD)
    c.alignment = _left()
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 10

    ws.merge_cells("B4:C4")
    c = ws["B4"]
    c.value = "Forensisch Statistisch Onderzoek — Ziekenhuis De Beemsterhof, Afdeling C-3"
    c.font = _font(False, 11, "555555", italic=True)
    c.alignment = _left()
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 10

    # Zaakinfo
    info = [
        ("Leerlingnummer",    str(leerlingnummer)),
        ("Ziekenhuis",        "De Beemsterhof"),
        ("Afdeling",          "C-3 Interne Geneeskunde"),
        ("Onderzoeksperiode", "8 januari 2024 – 30 maart 2024 (12 weken)"),
        ("Aantal diensten",   "252  (3 diensten × 84 dagen)"),
        ("Status",            "⚠  Actief forensisch onderzoek"),
    ]
    for i, (label, val) in enumerate(info):
        r = 6 + i
        ws.row_dimensions[r].height = 20
        c_l = ws.cell(r, 2, label)
        c_l.font = _font(True, 10)
        c_l.fill = _fill(K_GRIJS)
        c_l.border = _border()
        c_l.alignment = _left()
        c_v = ws.cell(r, 3, val)
        c_v.font = _font(False, 10)
        c_v.fill = _fill(K_WIT)
        c_v.border = _border()
        c_v.alignment = _left()

    ws.row_dimensions[12].height = 14

    # Weekopdrachten — beknopt overzicht passend bij de website
    weken = [
        ("WEEK 1 — Basisanalyse  (~90 min)", K_KOPTEKST, [
            "A  |  Dataset klaarmaken: maak zelf een Excel-tabel (Ctrl+T) → naam: Beemsterhof",
            "A  |  Activeer de totaalrij; sorteer en filter om opvallende diensten te ontdekken",
            "A  |  Pas de celopmaak aan en kies zelf een grenswaarde voor 'verhoogd risico'",
            "B  |  Nieuw tabblad 'Analyse': bereken gem./max/min patiënten met GEMIDDELDE, MAX, MIN",
            "B  |  Bereken gem. overlijdens per dienst met GEMIDDELDE.ALS + absolute verwijzingen ($)",
            "B  |  Bereken totaal overlijdens per arts met SOM.ALS en AFRONDEN",
            "B  |  Bereken mediaan, standaardafwijking, Q1, Q3 en uitbijtergrens van overlijdens",
            "→  Inleveren: Word-document (screenshots + toelichting) + Week1_[leerlingnummer].xlsx",
        ]),
        ("WEEK 2 — Verbanden & Draaitabellen  (~90 min)", K_ROOD, [
            "C  |  Kolom N: Sterftequotient — bedenk zelf de formule (overlijden ÷ patiënten)",
            "C  |  Kolom O: Risicocategorie — hercodeer met geneste ALS, kies zelf de grenzen",
            "C  |  Tabblad Analyse: gem. sterftequotiënt per arts met GEMIDDELDE.ALS",
            "D  |  Draaitabel 1: Arts × Dienst, waarden = gem. Sterftequotient (als %)",
            "D  |  Draaitabel 2: frequentietabel Risicocategorie",
            "D  |  Draaitabel 3: kruistabel Arts × Risicocategorie",
            "D  |  Tijdlijnfilter op Draaitabel 1: vergelijk eerste vs. alle 12 weken",
            "D  |  Draaitabel 4: toets argument De Wit (bezetting vs. quotiënt per arts)",
            "E  |  Groepeer Aantal_patienten in klassen via draaitabel (rechtsmuisknop → Groeperen)",
            "→  Inleveren: Word-document uitgebreid + Week2_[leerlingnummer].xlsx",
        ]),
        ("WEEK 3 — Visualisatie & Conclusie  (~90 min)", K_GROEN, [
            "F  |  Gegroepeerd staafdiagram op basis van Draaitabel 1",
            "F  |  Gecombineerd diagram met secundaire as: bezetting (staaf) vs. quotiënt (lijn)",
            "F  |  Cirkeldiagram van de Risicocategorie-verdeling",
            "G  |  Draaigrafiek (lijndiagram): overlijdens per datum, met tijdlijnfilter",
            "G  |  Groepeer datums op week in de draaigrafiek",
            "G  |  Voeg Arts toe als legenda aan de draaigrafiek",
            "H  |  Nieuw tabblad 'BoxplotData': vier kolommen (één per arts) via AutoFilter",
            "H  |  Boxplot van alle vier artsen + uitbijtergrens berekenen",
            "H  |  Vergelijk twee boxplots (verdachte vs. overige drie)",
            "H  |  Bereken mediaan per arts — vergelijk met gemiddelde",
            "I  |  Eindconclusie 150–250 woorden in Word (§1 Bevinding / §2 Bewijs / §3 Alternatief / §4 Oordeel)",
            "→  Inleveren: volledig Word-document + Rapport_Sterfgeval_[leerlingnummer].xlsx",
        ]),
    ]

    cur = 13
    for titel, kleur, taken in weken:
        ws.row_dimensions[cur].height = 26
        ws.merge_cells(f"B{cur}:C{cur}")
        c = ws.cell(cur, 2, titel)
        c.font = _font(True, 11, K_WIT)
        c.fill = _fill(kleur)
        c.alignment = _left()
        c.border = _border()
        cur += 1

        for taak in taken:
            ws.row_dimensions[cur].height = 17
            ws.merge_cells(f"B{cur}:C{cur}")
            c = ws.cell(cur, 2, taak)
            c.font = _font(False, 9)
            c.fill = _fill(K_WIT)
            c.alignment = _left()
            c.border = _border()
            cur += 1

        cur += 1  # lege rij tussen weken

    # Disclaimer
    ws.row_dimensions[cur].height = 14
    cur += 1
    ws.merge_cells(f"B{cur}:C{cur}")
    c = ws.cell(cur, 2,
        "⚠  Deze dataset is uniek voor dit leerlingnummer en fictief gegenereerd voor onderwijsdoeleinden.")
    c.font = _font(False, 9, "888888", italic=True)
    c.alignment = _left()

    ws.merge_cells(f"B{cur+1}:C{cur+1}")
    c2 = ws.cell(cur + 1, 2,
        "De data in tabblad 'Data' is fictief gegenereerd en uniek voor dit leerlingnummer.")
    c2.font = _font(False, 9, "888888", italic=True)
    c2.alignment = _left()


# ── Tabblad 2: Data ───────────────────────────────────────────────────────────

KOLOMMEN = [
    ("Datum",               14, "datum",      "DD-MM-YYYY", "center"),
    ("Dag",                 12, "dag",         None,         "left"),
    ("Weeknummer",          11, "weeknr",      "#,##0",      "center"),
    ("Dienst",              18, "dienst",      None,         "left"),
    ("Aanvangstijd",        12, "start",       None,         "center"),
    ("Arts",                24, "arts_naam",   None,         "left"),
    ("Arts_code",            9, "arts_code",   None,         "center"),
    ("Verpleegkundige",     18, "vplk",        None,         "left"),
    ("Aantal_patienten",    15, "n_pat",       "#,##0",      "center"),
    ("Ontslagen",           12, "ontslagen",   "#,##0",      "center"),
    ("Stabiel",             10, "stabiel",     "#,##0",      "center"),
    ("Overlijden",          12, "overlijden",  "#,##0",      "center"),
    ("Overlijdenskans_pct", 17, "kans_pct",    "0.00%",      "center"),
]

def _maak_data(wb, rijen):
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Koptekstrij — gestijld maar nog GEEN Excel-tabel (leerling maakt die zelf via A1)
    for i, (hdr, breedte, _, fmt, _align) in enumerate(KOLOMMEN):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = breedte
        c = ws.cell(1, col, hdr)
        _style_hdr(c)

    ws.row_dimensions[1].height = 28

    for r_idx, rij in enumerate(rijen):
        er = r_idx + 2
        bg = K_WIT if r_idx % 2 == 0 else K_GRIJS

        waarden = [
            rij["datum"], rij["dag"], rij["weeknr"],
            rij["dienst"], rij["start"],
            rij["arts_naam"], rij["arts_code"], rij["vplk"],
            rij["n_pat"], rij["ontslagen"], rij["stabiel"],
            rij["overlijden"], rij["kans_pct"] / 100,
        ]

        ws.row_dimensions[er].height = 16

        for c_idx, (_, _, _, fmt, align) in enumerate(KOLOMMEN):
            col = c_idx + 1
            c = ws.cell(er, col, waarden[c_idx])
            c.fill = _fill(bg)
            c.border = _border()
            c.font = _font(False, 9)
            c.alignment = _center() if align == "center" else _left()
            if fmt:
                c.number_format = fmt

    # Geen ws.add_table() — leerling doet dit zelf in opdracht A1
    return ws


# ── Publieke functie ──────────────────────────────────────────────────────────

def _stel_metadata_in(wb, leerlingnummer: str, lesgroep: str):
    """Stel document-eigenschappen in zodat het bestand herleidbaar is naar de leerling."""
    from openpyxl.packaging.core import DocumentProperties
    wb.properties = DocumentProperties()
    wb.properties.creator    = f"Operatie Sterfgeval — {lesgroep} — leerling {leerlingnummer}"
    wb.properties.title      = f"Dataset Beemsterhof C-3 [{leerlingnummer}]"
    wb.properties.subject    = "Forensisch Statistisch Onderzoek — HAVO/VWO Wiskunde"
    wb.properties.description = (
        f"Persoonlijke dataset gegenereerd voor leerlingnummer {leerlingnummer}, "
        f"lesgroep {lesgroep}. Dit bestand is uniek en herleidbaar. "
        "Inleveren als: Rapport_Sterfgeval_[leerlingnummer].xlsx"
    )
    wb.properties.keywords   = f"{leerlingnummer} {lesgroep} BeemsterhofC3"
    wb.properties.category   = "Wiskunde statistiek opdracht"


def genereer_xlsx_bytes(leerlingnummer: str, lesgroep: str = "onbekend") -> bytes:
    """Genereer een xlsx bestand als bytes-object (voor Streamlit download)."""
    rijen, verdachte_idx = _genereer_rijen(leerlingnummer)
    wb = Workbook()
    _maak_opdracht(wb, leerlingnummer, verdachte_idx)
    _maak_data(wb, rijen)
    _stel_metadata_in(wb, leerlingnummer, lesgroep)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def genereer_xlsx_bestand(leerlingnummer: str, lesgroep: str = "onbekend", pad: str = None) -> str:
    """Sla een xlsx bestand op schijf op (voor batch-gebruik)."""
    if pad is None:
        pad = f"Dataset_BeemsterhofC3_{leerlingnummer}.xlsx"
    rijen, verdachte_idx = _genereer_rijen(leerlingnummer)
    wb = Workbook()
    _maak_opdracht(wb, leerlingnummer, verdachte_idx)
    _maak_data(wb, rijen)
    _stel_metadata_in(wb, leerlingnummer, lesgroep)
    wb.save(pad)
    print(f"✅  {pad}  ({len(rijen)} rijen, verdachte: {ARTSEN[verdachte_idx]['naam']})")
    return pad


if __name__ == "__main__":
    import sys
    args = sys.argv[1:] or ["144555"]
    for ln in args:
        genereer_xlsx_bestand(ln)
